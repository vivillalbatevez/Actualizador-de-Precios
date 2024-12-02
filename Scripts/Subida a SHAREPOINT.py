#!/usr/bin/env python
# coding: utf-8

# In[134]:


import os
import io
import pandas as pd
from cryptography.fernet import Fernet
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.user_credential import UserCredential
from office365.sharepoint.files.file import File
import urllib.parse
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
import psutil


# # Rutas

# In[ ]:


# Rutas relativas para producto final (ajusta si es necesario en el futuro)
#Inputs
ruta_credenciales = os.path.join(os.path.dirname(__file__), "../Credenciales.txt")
Tabla_Precios_Path = os.path.join(os.path.dirname(__file__), "../Config/Tabla_Base.xlsx")


# #Credenciales
# ruta_credenciales = r"D:\MAXIMIA\PROYECTO ESTIMACIONES\Actualizador de precios\Credenciales.txt"
# Tabla_Precios_Path = r"D:\MAXIMIA\PROYECTO ESTIMACIONES\Actualizador de precios\Config\Tabla_Base.xlsx"

# # Funciones

# In[136]:


def leer_datos_configuracion(ruta):
    datos = {}
    with open(ruta, 'r', encoding='utf-8') as archivo:
        for linea in archivo:
            linea = linea.strip()
            if linea and '=' in linea:
                clave, valor = linea.split('=', 1)
                clave = clave.strip()
                valor = valor.strip()

                # Verificar y eliminar comillas si las hay
                if (valor.startswith('"') and valor.endswith('"')) or \
                   (valor.startswith("'") and valor.endswith("'")):
                    valor = valor[1:-1]  # Eliminar las comillas del inicio y fin

                datos[clave] = valor
    return datos


# In[137]:


def autenticarse_y_obtener_datos(f_fernet, usuario_enc, contraseña_enc, sitio_enc, ruta_archivo_enc, hoja_excel_enc):
    usuario = f_fernet.decrypt(usuario_enc).decode()
    contraseña = f_fernet.decrypt(contraseña_enc).decode()
    sitio = f_fernet.decrypt(sitio_enc).decode()
    ruta_archivo = f_fernet.decrypt(ruta_archivo_enc).decode()
    hoja_excel = f_fernet.decrypt(hoja_excel_enc).decode()
    ctx = ClientContext(sitio).with_credentials(UserCredential(usuario, contraseña))
    return ctx, ruta_archivo, hoja_excel


# # Configurar conexion

# In[138]:


# --- Generar una clave de encriptación en tiempo de ejecución ---
clave = Fernet.generate_key()
f = Fernet(clave)


# In[139]:


# --- almacenar autenticacion encriptada ---
datos = leer_datos_configuracion(ruta_credenciales)
usuario_encriptado = f.encrypt(datos.get('usuario').encode())
contraseña_encriptada = f.encrypt(datos.get('contraseña').encode())
sitio_encriptado = f.encrypt(datos.get('sitio').encode())
ruta_archivo_encriptada = f.encrypt(datos.get('ruta_archivo').encode())
hoja_excel_encriptada = f.encrypt(datos.get('hoja_excel').encode())


# # Autenticarse en SharePoint

# In[140]:


# Autenticarse y obtener el contexto de SharePoint y la ruta del archivo
ctx, ruta_archivo, hoja_excel = autenticarse_y_obtener_datos(
    f,
    usuario_encriptado,
    contraseña_encriptada,
    sitio_encriptado,
    ruta_archivo_encriptada,
    hoja_excel_encriptada
)


# # Carga de DF que seran subidos a Sharepoint

# In[141]:


output_tabla_principal = io.BytesIO()


# In[142]:


tabla_principal = pd.read_excel(Tabla_Precios_Path, sheet_name="Tabla_Base")
tabla_redet = pd.read_excel(Tabla_Precios_Path, sheet_name="Redeterminaciones")


# In[143]:


tabla_principal.head(5)


# In[144]:


tabla_redet.head(5)


# In[145]:


# Crea el archivo con múltiples hojas
with pd.ExcelWriter(output_tabla_principal, engine='openpyxl') as writer:
    tabla_principal.to_excel(writer, index=False, sheet_name='Tabla_Base')
    tabla_redet.to_excel(writer, index=False, sheet_name='Redeterminaciones')


# In[146]:


output_tabla_principal.seek(0)  # Regresa al inicio del archivo


# # Modificaciones esteticas

# In[147]:


wb = load_workbook(output_tabla_principal)


# In[148]:


# Iterar por las hojas y agregar formato de tabla
for sheet_name in ['Tabla_Base', 'Redeterminaciones']:  # Asegúrate de usar los nombres correctos
    ws = wb[sheet_name]
    
    # Determina el rango de la tabla (ajustar según tus datos)
    max_row = ws.max_row
    max_col = ws.max_column
    rango_tabla = f"A1:{ws.cell(row=max_row, column=max_col).coordinate}"
    
    # Crear la tabla
    tabla = Table(displayName=f"Tabla_{sheet_name}", ref=rango_tabla)
    
    # Agregar estilo a la tabla
    estilo = TableStyleInfo(
        name="TableStyleMedium9",  # Cambia por otro estilo si lo prefieres
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    tabla.tableStyleInfo = estilo
    ws.add_table(tabla)


# In[149]:


for ws in wb.worksheets:
    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width


# In[150]:


output_formatted = io.BytesIO()
wb.save(output_formatted)
output_formatted.seek(0)


# # Sobrescribir sharepoint

# In[151]:


# Extrae la ruta de la carpeta y el nombre del archivo de 'ruta_archivo'
ruta_carpeta, nombre_archivo = ruta_archivo.rsplit('/', 1)


# In[152]:


# Decodifica la ruta en caso de que contenga caracteres especiales
ruta_carpeta_decoded = urllib.parse.unquote(ruta_carpeta)
# Construye la URL relativa completa del archivo
ruta_archivo_completa = f"{ruta_carpeta}/{nombre_archivo}"


# In[153]:


# Sobrescribe el archivo en SharePoint
File.save_binary(
    ctx,
    ruta_archivo_completa,
    output_formatted
)

