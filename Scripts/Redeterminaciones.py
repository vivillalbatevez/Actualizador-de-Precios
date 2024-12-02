#!/usr/bin/env python
# coding: utf-8

# In[35]:


import pandas as pd
import numpy as np
from dateutil.relativedelta import relativedelta
from datetime import datetime
import xlsxwriter
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# In[ ]:


# Rutas relativas para producto final (ajusta si es necesario en el futuro)
#Inputs
input_user_file = os.path.join(os.path.dirname(__file__), "../Input/Input.xlsx")
tabla_base_path = os.path.join(os.path.dirname(__file__), "../Config/Tabla_Base.xlsx")

#Outputs
output_file_path = os.path.join(os.path.dirname(__file__), "../Output/tabla_Final.xlsx")


# # Ruta del archivo de inputs
# input_user_file = r"D:\MAXIMIA\PROYECTO ESTIMACIONES\Actualizador de precios\Input\Input.xlsx"
# tabla_base_path = r"D:\MAXIMIA\PROYECTO ESTIMACIONES\Actualizador de precios\Config\Tabla_Base.xlsx"
# 
# output_file_path = r"D:\MAXIMIA\PROYECTO ESTIMACIONES\Actualizador de precios\Output\tabla_Final.xlsx"

# In[37]:


fecha_act = pd.read_excel(input_user_file, sheet_name='Fechas a act')
redeter_input = pd.read_excel(input_user_file, sheet_name='Redeterminaciones')
redeter_base = pd.read_excel(tabla_base_path,sheet_name='Redeterminaciones')
tabla_final= pd.read_excel(output_file_path,sheet_name='tabla final')


# In[38]:


tabla_final


# In[39]:


tabla_final['Fecha'] = pd.to_datetime(tabla_final['Fecha'], dayfirst=True, errors='coerce')

# Inicializar la tabla de redeterminaciones como una copia de la tabla base
redeterminaciones = redeter_base.copy()
# Convertir la columna de fecha al formato dd/mm/yyyy
redeterminaciones['Fecha'] = pd.to_datetime(redeterminaciones['Fecha']).dt.strftime('%d/%m/%Y')


# In[40]:


def calcular_redeterminaciones(fecha_act, redeter_input, redeter_base, tabla_final):
    resultados = []

    # Convertir la columna 'Fechas' de fecha_act a datetime
    fecha_act['Fechas'] = pd.to_datetime(fecha_act['Fechas'], dayfirst=True)

    # Convertir la columna 'Fecha' de tabla_final y redeter_base a datetime en formato 'dd/mm/yyyy'
    tabla_final['Fecha'] = pd.to_datetime(tabla_final['Fecha'], dayfirst=True)
    redeter_base['Fecha'] = pd.to_datetime(redeter_base['Fecha'], dayfirst=True)

    # Tomar la fecha de la primera fila de fecha_act
    fecha = fecha_act['Fechas'].iloc[0]

    # Iterar sobre cada fila en redeter_input
    for _, row in redeter_input.iterrows():
        cliente = row['Codigo Cliente']
        nombre_cliente = row['Cliente']
        ccosto = row['Codigo Ccosto']
        nombre_ccosto = row['Ccosto']
        col_apoyo = row['Col  apoyo']

        # Obtener el coeficiente correspondiente a esa fecha en tabla_final
        coeficiente_row = tabla_final[
            (tabla_final['Codigo Cliente'] == cliente) & 
            (tabla_final['Cliente'] == nombre_cliente) & 
            (tabla_final['Codigo Ccosto'] == ccosto) & 
            (tabla_final['Fecha'].dt.to_period('M') == fecha.to_period('M'))
        ]
        
        if coeficiente_row.empty:
            
            continue  # Si no hay coeficiente, pasa al siguiente registro
        coeficiente = coeficiente_row['Coeficiente'].values[0]

        # Obtener el precio final del mes anterior de la tabla redeter_base
        fecha_mes_anterior = (fecha - relativedelta(months=1)).strftime('%d/%m/%Y')
        
        precio_anterior_row = redeter_base[
            (redeter_base['Codigo Cliente'] == cliente) & 
            (redeter_base['Cliente'] == nombre_cliente) & 
            (redeter_base['Codigo Ccosto'] == ccosto) & 
            (redeter_base['Fecha'] == pd.to_datetime(fecha_mes_anterior, dayfirst=True))
        ]

        if precio_anterior_row.empty:
            
            continue

        precio_final_anterior = precio_anterior_row['Precio Final'].values[0]
        precio_base = precio_anterior_row['Precio base'].values[0]
        articulo = precio_anterior_row['Cod Articulo'].values[0]
        nombre_articulo = precio_anterior_row['Articulo'].values[0]

        
        # Calcular el precio final actualizado y la redeterminación
        precio_final_actualizado = precio_final_anterior * coeficiente
        redeterminacion = precio_final_actualizado - precio_base

        # Crear diccionario de resultados para cada fila
        resultado = {
            'Codigo Cliente': cliente,
            'Cliente': nombre_cliente,
            'Codigo Ccosto': ccosto,
            'Ccosto': nombre_ccosto,
            'Cod Articulo': articulo,
            'Articulo': nombre_articulo,
            'Fecha': fecha.strftime('%d/%m/%Y'),
            'Precio base': float(precio_base),
            'Precio Final': precio_final_actualizado,
            'Redeterminacion': redeterminacion,
            'Col  apoyo': col_apoyo
        }
        resultados.append(resultado)

    # Convertir lista de resultados en un DataFrame y retornar
    df_resultados = pd.DataFrame(resultados)
    # Formatear la columna 'Fecha' a %d/%m/%y
    if 'Fecha' in tabla_final.columns:
        tabla_final['Fecha'] = pd.to_datetime(tabla_final['Fecha'], dayfirst=True).dt.strftime('%d/%m/%Y')
    
    return df_resultados


# In[41]:


# Llamar a la función y obtener los nuevos datos de redeterminaciones
nuevas_redeterminaciones = calcular_redeterminaciones(fecha_act, redeter_input, redeter_base, tabla_final)

# Concatenar las nuevas redeterminaciones a la tabla redeterminaciones existente
tabla_redeterminaciones_actualizada = pd.concat([redeterminaciones, nuevas_redeterminaciones], ignore_index=True)

tabla_final['SinActPorGatillo'] = tabla_final['SinActPorGatillo'].astype(str)


# In[42]:


with pd.ExcelWriter(output_file_path, engine='xlsxwriter') as writer:
    tabla_final.to_excel(writer, sheet_name='Tabla Final', index=False)
    tabla_redeterminaciones_actualizada.to_excel(writer, sheet_name='Redeterminaciones', index=False)


# # Modificaciones Esteticas

# In[43]:


def ajustar_ancho_columnas(ws):
    #Ajusta el ancho de las columnas en una hoja
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width


# In[44]:


def crear_tabla_y_frizar(ws, contador):
    # Contador de tablas específico para cada hoja
    table_counter = contador
    
    # Obtener el nombre de la hoja automáticamente
    sheet_name = ws.title
    
    # Determinar el rango de la tabla
    min_col, min_row = 1, 1
    max_col = ws.max_column
    max_row = ws.max_row
    table_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    
    # Crear el nombre único de la tabla para esta hoja
    table_name = f"Table_{sheet_name.replace(' ', '_')}_{table_counter}"
    table_counter += 1  # Incrementar el contador solo para tablas adicionales en la misma hoja

    # Crear y estilizar la tabla
    table = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    
    # Añadir la tabla a la hoja de trabajo
    try:
        ws.add_table(table)
    except ValueError as e:
        print(f"Error adding table to sheet {sheet_name}: {e}")
    
    # Fijar la primera fila (encabezados)
    ws.freeze_panes = ws['A2']


# In[45]:


wbTablaFinal = load_workbook(output_file_path)


# In[46]:


# Configurar el workbook para recalcular automáticamente las fórmulas al abrir el archivo
wbTablaFinal.calcMode = 'auto'


# In[47]:


# Ajustar el ancho de las columnas de todas las hojas
contador = 1
for sheet in wbTablaFinal.sheetnames:
    ws = wbTablaFinal[sheet]
    ajustar_ancho_columnas(ws)
    crear_tabla_y_frizar(ws, contador)
    contador = contador + 1


# In[48]:


# Guardar el archivo con los cambios
wbTablaFinal.save(output_file_path)
wbTablaFinal.close()
del wbTablaFinal

