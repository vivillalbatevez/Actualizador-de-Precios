#!/usr/bin/env python
# coding: utf-8

# In[36]:


import pandas as pd
import warnings
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import os


# In[37]:


# Suprime las advertencias
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# ### Definimos rutas y creamos los DF con pandas

# In[ ]:


# Rutas relativas para producto final (ajusta si es necesario en el futuro)
#Inputs
ruta_Porcentaje_ACT_Polinomicas = os.path.join(os.path.dirname(__file__), "../Config/Porcentaje_ACT.xlsx")
ruta_Tabla_Base_Precios = os.path.join(os.path.dirname(__file__), "../Config/Tabla_Base.xlsx")
ruta_fechas = os.path.join(os.path.dirname(__file__), "../Input/Input.xlsx")

#Outputs
ruta_guardado_tabla_final = os.path.join(os.path.dirname(__file__), "../Output/Para_Ges.xlsx")


# # Inputs
# ruta_Porcentaje_ACT_Polinomicas = r"D:\MAXIMIA\PROYECTO ESTIMACIONES\Actualizador de precios\Config\Porcentaje_ACT.xlsx"
# ruta_Tabla_Base_Precios = r"D:\MAXIMIA\PROYECTO ESTIMACIONES\Actualizador de precios\Config\Tabla_Base.xlsx"
# ruta_fechas = r"D:\MAXIMIA\PROYECTO ESTIMACIONES\Actualizador de precios\Input\Input.xlsx"
# 
# # Outputs
# ruta_guardado_tabla_final = r"D:\MAXIMIA\PROYECTO ESTIMACIONES\Actualizador de precios\Output\Para_Ges.xlsx"

# In[39]:


# Lee los archivos de Excel
Tabla_Porcentaje_ACT_Polinomicas = pd.read_excel(ruta_Porcentaje_ACT_Polinomicas, sheet_name="Principal")

Tabla_Input_Fechas = pd.read_excel(ruta_fechas, sheet_name="Fechas a act")

Tabla_Tabla_Base_Precios = pd.read_excel(ruta_Tabla_Base_Precios, sheet_name="Tabla_Base")


# ### Creacion tabla articulos * contrato

# In[40]:


# Concatenar las columnas deseadas con un guion como separador y renombrar directamente a 'Identificador'
Tabla_Porcentaje_ACT_Polinomicas["Identificador"] = Tabla_Porcentaje_ACT_Polinomicas["Cod cliente"].astype(str) + '-/-/-' + Tabla_Porcentaje_ACT_Polinomicas["Cliente"].astype(str) + '-/-/-' + Tabla_Porcentaje_ACT_Polinomicas["Codigo CC"].astype(str) + '-/-/-' + Tabla_Porcentaje_ACT_Polinomicas["Ccosto"].astype(str) + '-/-/-' + Tabla_Porcentaje_ACT_Polinomicas["Col apoyo"].astype(str)

Tabla_Porcentaje_ACT_Polinomicas.head(3)


# In[41]:


# Selecciona solo las columnas 'Identificador', 'Cod Articulo', y 'Articulo' en la tabla final
Tabla_Porcentaje_ACT_Polinomicas = Tabla_Porcentaje_ACT_Polinomicas[["Identificador", "Cod Articulo", "Articulo"]]

# Muestra las primeras filas para verificar el resultado
Tabla_Porcentaje_ACT_Polinomicas.tail(5)


# ### Creacion tabla Precios act

# In[42]:


Tabla_Tabla_Base_Precios['Fecha'] = pd.to_datetime(Tabla_Tabla_Base_Precios['Fecha'], format='%d/%m/%Y')


# In[43]:


Tabla_Tabla_Base_Precios.head(3)


# In[44]:


Tabla_Input_Fechas


# In[45]:


#Filtro de fechas
# Convertir la columna de fechas a una lista
lista_fechas = Tabla_Input_Fechas['Fechas'].tolist() 

# Filtramos para conservar solo las filas donde la fecha esté en la lista de fechas
Tabla_Tabla_Base_Precios = Tabla_Tabla_Base_Precios[Tabla_Tabla_Base_Precios['Fecha'].isin(lista_fechas)]


# In[46]:


Tabla_Tabla_Base_Precios.head(3)


# In[47]:


Tabla_Tabla_Base_Precios["Identificador"] = Tabla_Tabla_Base_Precios["Codigo Cliente"].astype(str) + '-/-/-' + Tabla_Tabla_Base_Precios["Cliente"].astype(str) + '-/-/-' + Tabla_Tabla_Base_Precios["Codigo Ccosto"].astype(str) + '-/-/-' + Tabla_Tabla_Base_Precios["Ccosto"].astype(str) + '-/-/-' + Tabla_Tabla_Base_Precios["Col  apoyo"].astype(str)
    
# Crear tabla2 con las columnas seleccionadas
Tabla_Tabla_Base_Precios = Tabla_Tabla_Base_Precios[["Identificador", "Precio"]]

Tabla_Tabla_Base_Precios.head(5)


# ### Creacion Tabla Final

# In[48]:


# Realizar la unión de las dos tablas usando la columna 'Identificador' como clave para generar tabla_final
tabla_final = pd.merge(Tabla_Porcentaje_ACT_Polinomicas, Tabla_Tabla_Base_Precios, on="Identificador", how="inner")


# In[49]:


tabla_final


# In[50]:


# Separar la columna 'Identificador' en partes y asignar directamente a nuevas columnas en tabla_final
tabla_final[['Codigo Cliente', 'Cliente', 'Codigo C.costo', 'Ccosto', 'Col apoyo']] = tabla_final['Identificador'].str.split('-/-/-', expand=True)
tabla_final.head(3)


# In[51]:


# Seleccionar y reorganizar las columnas en la tabla final
tabla_final = tabla_final[['Codigo Cliente', 'Cliente', 'Codigo C.costo', 'Ccosto', 'Col apoyo', 'Cod Articulo', 'Articulo', 'Precio']]
tabla_final.head(3)


# ### Guardamos Excel

# In[52]:


# Guardar la tabla final en un archivo Excel (sobrescribiendo cualquier archivo anterior)
tabla_final.to_excel(ruta_guardado_tabla_final, index=False)

print(f"Archivo guardado exitosamente en {ruta_guardado_tabla_final}")


# # Modificaciones esteticas

# #### Funciones

# In[53]:


def ajustar_ancho_columnas(ws):
    #Ajusta el ancho de las columnas en una hoja
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width


# In[54]:


def crear_tabla_y_frizar(ws, contador):
    # Contador de tablas específico para cada hoja
    table_counter = contador
    
    # Obtener el nombre de la hoja automáticamente
    sheet_name = ws.title
    
    # Determinar el rango de la tabla
    min_col, min_row = 1, 1
    max_col = ws.max_column
    max_row = ws.max_row
    table_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    
    # Crear el nombre único de la tabla para esta hoja
    table_name = f"Table_{sheet_name.replace(' ', '_')}_{table_counter}"
    table_counter += 1  # Incrementar el contador solo para tablas adicionales en la misma hoja

    # Crear y estilizar la tabla
    table = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    
    # Añadir la tabla a la hoja de trabajo
    try:
        ws.add_table(table)
    except ValueError as e:
        print(f"Error adding table to sheet {sheet_name}: {e}")
    
    # Fijar la primera fila (encabezados)
    ws.freeze_panes = ws['A2']


# #### Fin funciones 

# In[55]:


wb = load_workbook(ruta_guardado_tabla_final)


# In[56]:


contador = 1
for sheet in wb.sheetnames:
    ws = wb[sheet]
    ajustar_ancho_columnas(ws)
    crear_tabla_y_frizar(ws, contador)
    contador = contador + 1


# In[57]:


# Guardar el archivo con los cambios
wb.save(ruta_guardado_tabla_final)
wb.close()

