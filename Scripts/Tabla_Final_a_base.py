#!/usr/bin/env python
# coding: utf-8

# In[114]:


import pandas as pd
import warnings
import xlsxwriter
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import xlwings as xw
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
import sys


# In[ ]:


# Rutas relativas para producto final (ajusta si es necesario en el futuro)

input_user_file = os.path.join(os.path.dirname(__file__), "../Input/Input.xlsx")
output_file_path = os.path.join(os.path.dirname(__file__), "../Output/tabla_Final.xlsx")
tabla_base_path = os.path.join(os.path.dirname(__file__), "../Config/Tabla_Base.xlsx")


# # Ruta del archivo de inputs
# input_user_file = r"D:\MAXIMIA\PROYECTO ESTIMACIONES\Actualizador de precios\Input\Input.xlsx"
# output_file_path = r"D:\MAXIMIA\PROYECTO ESTIMACIONES\Actualizador de precios\Output\tabla_Final.xlsx"
# tabla_base_path = r"D:\MAXIMIA\PROYECTO ESTIMACIONES\Actualizador de precios\Config\Tabla_Base.xlsx"

# In[116]:


tabla_final= pd.read_excel(output_file_path,sheet_name='Tabla Final')
redeterminaciones = pd.read_excel(output_file_path,sheet_name='Redeterminaciones')
fecha_act = pd.read_excel(input_user_file, sheet_name='Fechas a act')
tabla_final['SinActPorGatillo'] = tabla_final['SinActPorGatillo'].astype(str)


# In[117]:


redeterminaciones


# In[118]:


with pd.ExcelWriter(tabla_base_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    # Reemplaza la hoja 'tabla_base' con los datos de 'tabla_final'
    tabla_final.to_excel(writer, sheet_name="Tabla_Base", index=False)

    # Reemplaza la hoja 'redeterminaciones' con los datos de 'redeterminaciones'
    redeterminaciones.to_excel(writer, sheet_name="Redeterminaciones", index=False)


# In[119]:


# Convertir la columna 'Fechas' de `fecha_act` y las columnas de fecha en las otras tablas al mismo formato
fecha_act['Fechas'] = pd.to_datetime(fecha_act['Fechas'], dayfirst=True)
tabla_final['Fecha'] = pd.to_datetime(tabla_final['Fecha'], dayfirst=True)
redeterminaciones['Fecha'] = pd.to_datetime(redeterminaciones['Fecha'], dayfirst=True)

# Obtener las fechas únicas de `fecha_act`
fechas_a_conservar = fecha_act['Fechas'].unique()

# Filtrar fechas en ambas tablas
tabla_final_filtrada = tabla_final[tabla_final['Fecha'].isin(fechas_a_conservar)].copy()
tabla_final_filtrada = tabla_final_filtrada.drop(['Cod Articulo', 'Articulo'], axis=1)
redeterminaciones_filtrada = redeterminaciones[redeterminaciones['Fecha'].isin(fechas_a_conservar)].copy()

# Convertir a formato de texto solo después de filtrar
tabla_final_filtrada['Fecha'] = tabla_final_filtrada['Fecha'].dt.strftime('%d/%m/%Y')
redeterminaciones_filtrada['Fecha'] = redeterminaciones_filtrada['Fecha'].dt.strftime('%d/%m/%Y')


# In[120]:


# Guardar los resultados en nuevas hojas en el archivo de salida
with pd.ExcelWriter(output_file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    tabla_final_filtrada.to_excel(writer, sheet_name="Tabla Final", index=False)
    redeterminaciones_filtrada.to_excel(writer, sheet_name="Redeterminaciones", index=False)


# # Funciones para mejoras esteticas

# In[121]:


def ajustar_ancho_columnas(ws):
    #Ajusta el ancho de las columnas en una hoja
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width


# In[122]:


def crear_tabla_y_frizar(ws, contador):
    # Contador de tablas específico para cada hoja
    table_counter = contador
    
    # Obtener el nombre de la hoja automáticamente
    sheet_name = ws.title
    
    # Determinar el rango de la tabla
    min_col, min_row = 1, 1
    max_col = ws.max_column
    max_row = ws.max_row
    table_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    
    # Crear el nombre único de la tabla para esta hoja
    table_name = f"Table_{sheet_name.replace(' ', '_')}_{table_counter}"
    table_counter += 1  # Incrementar el contador solo para tablas adicionales en la misma hoja

    # Crear y estilizar la tabla
    table = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    
    # Añadir la tabla a la hoja de trabajo
    try:
        ws.add_table(table)
    except ValueError as e:
        print(f"Error adding table to sheet {sheet_name}: {e}")
    
    # Fijar la primera fila (encabezados)
    ws.freeze_panes = ws['A2']


# # Mejoras esteticas tabla Final

# In[123]:


wbTablaFinal = load_workbook(output_file_path)


# In[124]:


# Configurar el workbook para recalcular automáticamente las fórmulas al abrir el archivo
wbTablaFinal.calcMode = 'auto'


# In[125]:


# Ajustar el ancho de las columnas de todas las hojas
contador = 1
for sheet in wbTablaFinal.sheetnames:
    ws = wbTablaFinal[sheet]
    ajustar_ancho_columnas(ws)
    crear_tabla_y_frizar(ws, contador)
    contador = contador + 1


# In[126]:


# Guardar el archivo con los cambios
wbTablaFinal.save(output_file_path)
wbTablaFinal.close()
del wbTablaFinal


# # Mejoras esteticas tabla Base

# In[127]:


wbPorcentaje = load_workbook(tabla_base_path)
# Configurar el workbook para recalcular automáticamente las fórmulas al abrir el archivo
wbPorcentaje.calcMode = 'auto'


# In[128]:


#Convertir a formato fecha corta la columna que yo le diga
ws_TablaBase = wbPorcentaje['Tabla_Base']
for celda in ws_TablaBase['G']:
    celda.number_format = 'DD/MM/YYYY'
    
#Convertir a formato fecha corta la columna que yo le diga
ws_Redeterminaciones = wbPorcentaje['Redeterminaciones']
for celda in ws_Redeterminaciones['G']:
    celda.number_format = 'DD/MM/YYYY'


# In[129]:


# Ajustar el ancho de las columnas de todas las hojas
contador = 1
wsTabla_base = wbPorcentaje['Tabla_Base']
ajustar_ancho_columnas(wsTabla_base)
crear_tabla_y_frizar(wsTabla_base, contador)
contador = contador + 1

wsTabla_redet = wbPorcentaje['Redeterminaciones']
ajustar_ancho_columnas(wsTabla_redet)
crear_tabla_y_frizar(wsTabla_redet, contador)
contador = contador + 1


# In[130]:


# Guardar el archivo con los cambios
wbPorcentaje.save(tabla_base_path)
wbPorcentaje.close()
del wbPorcentaje


# In[131]:


# Abrir el libro de Excel usando xlwings
app = xw.App(visible=False)  # Mantener Excel invisible durante la operación
wbPorcentaje2 = xw.Book(tabla_base_path)  # Abrir el archivo
wbPorcentaje2_Hoja_Principal = wbPorcentaje2.sheets["Tabla_Base"]  # Seleccionar la hoja
wbPorcentaje2_Hoja_Redet = wbPorcentaje2.sheets["Redeterminaciones"]  # Seleccionar la hoja


# In[132]:


# Encontrar la última fila con datos en las columnas de hoja principal
Prin_COLA_ultima_fila = wbPorcentaje2_Hoja_Principal.range('A2').end('down').row
Prin_COLC_ultima_fila = wbPorcentaje2_Hoja_Principal.range('C2').end('down').row
Prin_COLE_ultima_fila = wbPorcentaje2_Hoja_Principal.range('E2').end('down').row


# In[133]:


# Aplicar la fórmula en las filas de la columna B
for fila in range(2, Prin_COLA_ultima_fila + 1):
    formula = '=IFERROR(INDEX(Table_MaestroClientes_4[Codigo cliente], MATCH([@Cliente], Table_MaestroClientes_4[Cliente], 0)), "")'
    wbPorcentaje2_Hoja_Principal.range(f'A{fila}').formula = formula
    
# Aplicar la fórmula en las filas de la columna C
for fila in range(2, Prin_COLC_ultima_fila + 1):
    formula = '=IFERROR(INDEX(Table_MaestroCcostos_3[Codigo CC], MATCH([@Ccosto], Table_MaestroCcostos_3[Ccosto], 0)), "")'
    wbPorcentaje2_Hoja_Principal.range(f'C{fila}').formula = formula
    
# Aplicar la fórmula en las filas de la columna C
for fila in range(2, Prin_COLE_ultima_fila + 1):
    formula = '=IFERROR(INDEX(Table_Maestro_art_5[CODIGO], MATCH([@Articulo], Table_Maestro_art_5[ARTICULO], 0)), "")'
    wbPorcentaje2_Hoja_Principal.range(f'E{fila}').formula = formula


# In[134]:


# Encontrar la última fila con datos en las columnas de hoja principal
Redet_COLA_ultima_fila = wbPorcentaje2_Hoja_Redet.range('A2').end('down').row
Redet_COLC_ultima_fila = wbPorcentaje2_Hoja_Redet.range('C2').end('down').row
redet_COLE_ultima_fila = wbPorcentaje2_Hoja_Redet.range('E2').end('down').row


# In[135]:


# Aplicar la fórmula en las filas de la columna B
for fila in range(2, Redet_COLA_ultima_fila + 1):
    formula = '=IFERROR(INDEX(Table_MaestroClientes_4[Codigo cliente], MATCH([@Cliente], Table_MaestroClientes_4[Cliente], 0)), "")'
    wbPorcentaje2_Hoja_Redet.range(f'A{fila}').formula = formula
    
# Aplicar la fórmula en las filas de la columna C
for fila in range(2, Redet_COLC_ultima_fila + 1):
    formula = '=IFERROR(INDEX(Table_MaestroCcostos_3[Codigo CC], MATCH([@Ccosto], Table_MaestroCcostos_3[Ccosto], 0)), "")'
    wbPorcentaje2_Hoja_Redet.range(f'C{fila}').formula = formula
    
# Aplicar la fórmula en las filas de la columna E
for fila in range(2, redet_COLE_ultima_fila + 1):
    formula = '=IFERROR(INDEX(Table_Maestro_art_5[CODIGO], MATCH([@Articulo], Table_Maestro_art_5[ARTICULO], 0)), "")'
    wbPorcentaje2_Hoja_Redet.range(f'E{fila}').formula = formula


# In[136]:


# Definir el rango de la columna A (desde la fila 2 hasta la fila 1048576)
COLB_rango_validacion = wbPorcentaje2_Hoja_Principal.range('B2:B1048576')
COLD_rango_validacion = wbPorcentaje2_Hoja_Principal.range('D2:D1048576')
COLF_rango_validacion = wbPorcentaje2_Hoja_Principal.range('F2:F1048576')


# In[137]:


# Eliminar cualquier validación anterior
COLB_rango_validacion.api.Validation.Delete()
# Aplicar la validación de lista
COLB_rango_validacion.api.Validation.Add(
    3,  # xlValidateList (tipo de validación de lista)
    1,  # xlValidAlertStop (alerta de detención en caso de valor incorrecto)
    1,  # xlBetween (operador)
    "='MaestroClientes'!$B$2:$B$1048576"  # La referencia de la lista
)

# Eliminar cualquier validación anterior
COLD_rango_validacion.api.Validation.Delete()
# Aplicar la validación de lista
COLD_rango_validacion.api.Validation.Add(
    3,  # xlValidateList (tipo de validación de lista)
    1,  # xlValidAlertStop (alerta de detención en caso de valor incorrecto)
    1,  # xlBetween (operador)
    "='MaestroCcostos'!$B$2:$B$1048576"  # La referencia de la lista
)

# Eliminar cualquier validación anterior
COLF_rango_validacion.api.Validation.Delete()
# Aplicar la validación de lista
COLF_rango_validacion.api.Validation.Add(
    3,  # xlValidateList (tipo de validación de lista)
    1,  # xlValidAlertStop (alerta de detención en caso de valor incorrecto)
    1,  # xlBetween (operador)
    "='Maestro art'!$B$2:$B$1048576"  # La referencia de la lista
)


# In[138]:


# Definir el rango de la columna A (desde la fila 2 hasta la fila 1048576)
COLB_rango_validacion = wbPorcentaje2_Hoja_Redet.range('B2:B1048576')
COLD_rango_validacion = wbPorcentaje2_Hoja_Redet.range('D2:D1048576')
COLF_rango_validacion = wbPorcentaje2_Hoja_Redet.range('F2:F1048576')


# In[139]:


# Eliminar cualquier validación anterior
COLB_rango_validacion.api.Validation.Delete()
# Aplicar la validación de lista
COLB_rango_validacion.api.Validation.Add(
    3,  # xlValidateList (tipo de validación de lista)
    1,  # xlValidAlertStop (alerta de detención en caso de valor incorrecto)
    1,  # xlBetween (operador)
    "='MaestroClientes'!$B$2:$B$1048576"  # La referencia de la lista
)

# Eliminar cualquier validación anterior
COLD_rango_validacion.api.Validation.Delete()
# Aplicar la validación de lista
COLD_rango_validacion.api.Validation.Add(
    3,  # xlValidateList (tipo de validación de lista)
    1,  # xlValidAlertStop (alerta de detención en caso de valor incorrecto)
    1,  # xlBetween (operador)
    "='MaestroCcostos'!$B$2:$B$1048576"  # La referencia de la lista
)

# Eliminar cualquier validación anterior
COLF_rango_validacion.api.Validation.Delete()
# Aplicar la validación de lista
COLF_rango_validacion.api.Validation.Add(
    3,  # xlValidateList (tipo de validación de lista)
    1,  # xlValidAlertStop (alerta de detención en caso de valor incorrecto)
    1,  # xlBetween (operador)
    "='Maestro art'!$B$2:$B$1048576"  # La referencia de la lista
)


# In[140]:


# Guardar el archivo y cerrar Excel
wbPorcentaje2.save(tabla_base_path)
wbPorcentaje2.close()
app.quit()


# In[141]:


# Abrir el libro de Excel usando xlwings
app = xw.App(visible=False)  # Mantener Excel invisible durante la operación
wbPorcentaje2 = xw.Book(tabla_base_path)  # Abrir el archivo


# In[142]:


nombre_hoja = 'Tabla_Base'  # Cambia esto por el nombre de tu hoja
nombre_columna = 'J'  # Cambia esto por la letra de la columna que deseas modificar
hoja = wbPorcentaje2.sheets[nombre_hoja]

# Obtener el rango de celdas usado en la columna específica
rango_columna = hoja.range(f'{nombre_columna}1').expand('down')


# In[143]:


# Iterar solo en las filas de la columna específica para cambiar True a "True" y False a "False" como texto
for celda in rango_columna:
    if celda.value is True:
        celda.api.NumberFormat = "@"  # Aplicar formato de texto explícito
        celda.value = "True"
    elif celda.value is False:
        celda.api.NumberFormat = "@"  # Aplicar formato de texto explícito
        celda.value = "False"


# In[144]:


# Guardar el archivo y cerrar Excel
wbPorcentaje2.save(tabla_base_path)
wbPorcentaje2.close()
app.quit()

